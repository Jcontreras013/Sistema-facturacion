"""Importación de inventario desde archivos de otros sistemas (CSV/Excel).

No adivina "cualquier" archivo con certeza (eso no existe), pero cubre la
gran mayoría de exportaciones reales de sistemas de cobro:
- Detecta automáticamente cuál columna es cuál por el nombre del encabezado
  (código, nombre, precio, existencias, etc. en español o inglés).
- Si no puede adivinar, o adivina mal, el usuario corrige el mapeo a mano
  en la pantalla de previsualización antes de confirmar.
- Tolera CSV con distintos separadores/codificaciones y Excel viejo (.xls)
  o nuevo (.xlsx), además de precios con coma o punto decimal.
"""

import csv
import datetime
import io
import re
import unicodedata
import uuid
from decimal import Decimal, InvalidOperation

PRODUCT_FIELDS = [
    ("ignore", "Ignorar esta columna"),
    ("code", "Código / SKU"),
    ("barcode", "Código de barras"),
    ("name", "Nombre (obligatorio)"),
    ("category", "Categoría"),
    ("provider", "Proveedor"),
    ("unit", "Unidad de medida"),
    ("purchase_price", "Precio de compra"),
    ("sale_price", "Precio de venta"),
    ("tax_rate", "ISV / Impuesto (%)"),
    ("stock", "Existencias"),
    ("min_stock", "Stock mínimo"),
    ("expiration_date", "Fecha de vencimiento"),
]

FIELD_SYNONYMS = {
    "code": ["codigo", "cod", "sku", "clave", "id", "codigo producto", "codigo interno", "referencia", "ref"],
    "barcode": ["codigo de barras", "cod barras", "barras", "ean", "ean13", "upc", "codigo barra"],
    "name": ["nombre", "descripcion", "producto", "articulo", "detalle", "nombre producto", "descripcion articulo", "concepto"],
    "category": ["categoria", "familia", "departamento", "linea", "grupo", "rubro"],
    "provider": ["proveedor", "marca", "fabricante", "distribuidor"],
    "unit": ["unidad", "medida", "um", "u m", "presentacion"],
    "purchase_price": ["costo", "precio costo", "precio compra", "costo unitario", "precio de compra", "costo unit"],
    "sale_price": ["precio", "precio venta", "pvp", "precio de venta", "precio unitario", "venta", "precio publico"],
    "tax_rate": ["isv", "iva", "impuesto", "tasa impuesto", "tasa isv"],
    "stock": ["existencia", "existencias", "stock", "cantidad", "inventario", "cant", "saldo", "existencia actual"],
    "min_stock": ["stock minimo", "minimo", "existencia minima", "punto reorden", "stock de seguridad"],
    "expiration_date": ["vencimiento", "fecha vencimiento", "caducidad", "fecha caducidad", "fecha expiracion"],
}

UNIT_SYNONYMS = {
    "unidad": "unidad", "unid": "unidad", "und": "unidad", "u": "unidad", "pza": "unidad", "pieza": "unidad", "pzas": "unidad",
    "lb": "lb", "libra": "lb", "libras": "lb",
    "kg": "kg", "kilo": "kg", "kilogramo": "kg", "kilogramos": "kg",
    "caja": "caja", "cja": "caja", "cajas": "caja",
    "litro": "litro", "litros": "litro", "lt": "litro", "lts": "litro", "l": "litro",
    "paquete": "paquete", "paq": "paquete", "pqt": "paquete", "paquetes": "paquete",
}

DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%y", "%Y/%m/%d"]


def normalize_header(text):
    text = str(text or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def detect_column_mapping(headers):
    """Devuelve {campo: indice_columna} para los campos que logró adivinar."""
    normalized = [normalize_header(h) for h in headers]
    mapping = {}
    used = set()

    # 1) coincidencia exacta con algún sinónimo
    for field, synonyms in FIELD_SYNONYMS.items():
        for idx, h in enumerate(normalized):
            if idx in used or not h:
                continue
            if h in synonyms:
                mapping[field] = idx
                used.add(idx)
                break

    # 2) coincidencia parcial (contiene / está contenido) para lo que falte
    for field, synonyms in FIELD_SYNONYMS.items():
        if field in mapping:
            continue
        for idx, h in enumerate(normalized):
            if idx in used or not h:
                continue
            if any(syn and (syn in h or h in syn) for syn in synonyms):
                mapping[field] = idx
                used.add(idx)
                break

    return mapping


def _cell_to_str(cell):
    if cell is None:
        return ""
    if isinstance(cell, (datetime.datetime, datetime.date)):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, float):
        if cell.is_integer():
            return str(int(cell))
        return repr(cell)
    return str(cell)


def _parse_csv(fileobj):
    raw = fileobj.read()
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = [row for row in reader if any((cell or "").strip() for cell in row)]
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _parse_xlsx(fileobj):
    from openpyxl import load_workbook

    wb = load_workbook(fileobj, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(values_only=True):
        str_row = [_cell_to_str(c) for c in row]
        if any(cell.strip() for cell in str_row):
            rows.append(str_row)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _parse_xls(fileobj):
    import xlrd

    book = xlrd.open_workbook(file_contents=fileobj.read())
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        raw_row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        str_row = [_cell_to_str(c) for c in raw_row]
        if any(cell.strip() for cell in str_row):
            rows.append(str_row)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def parse_uploaded_file(fileobj, filename):
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("xlsx", "xlsm"):
        return _parse_xlsx(fileobj)
    if ext == "xls":
        return _parse_xls(fileobj)
    return _parse_csv(fileobj)


def parse_decimal(value, default=Decimal("0")):
    if value is None:
        return default
    s = str(value).strip()
    if not s:
        return default
    s = re.sub(r"[^0-9.,-]", "", s)
    if not s:
        return default
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        if re.match(r"^-?\d+,\d{1,2}$", s):
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return default


def parse_tax_rate(value):
    if value is None or not str(value).strip():
        return Decimal("15.00")
    s = str(value).replace("%", "").strip()
    d = parse_decimal(s, default=None)
    if d is None:
        return Decimal("15.00")
    if d <= 1:
        d = d * 100
    return d


def normalize_unit(value):
    key = normalize_header(value)
    return UNIT_SYNONYMS.get(key, "unidad")


def parse_date_flexible(value):
    s = str(value or "").strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def generate_code():
    return f"IMP-{uuid.uuid4().hex[:8].upper()}"
