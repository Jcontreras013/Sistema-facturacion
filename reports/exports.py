from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse


def xlsx_response(filename, headers, rows, sheet_title="Reporte"):
    """Genera un archivo .xlsx descargable a partir de encabezados y filas de datos."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="14804A", end_color="14804A", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        lengths = [len(str(header))] + [len(str(row[col_idx - 1])) for row in rows]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(lengths) + 4, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
